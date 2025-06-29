# Copyright (c) 2024, Eng. Saeed Al-Adawi and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe import _
import os
import json
import hashlib
import time
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
import pandas as pd
import pdfkit
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import csv
import io
import base64
from PIL import Image, ImageOps
import tempfile


class ReportExportUtility(Document):
    # pylint: disable=no-member
    # Frappe framework dynamically adds DocType fields to Document class

    def validate(self):
        """Validate export configuration before saving"""
        self.validate_export_settings()
        self.validate_arabic_settings()
        self.validate_mobile_settings()
        self.set_default_values()

    def before_save(self):
        """Set calculated fields before saving"""
        if not self.created_by:
            self.created_by = frappe.session.user

        if not self.expiry_date and self.retention_days:
            self.expiry_date = datetime.now() + timedelta(days=self.retention_days)

    def validate_export_settings(self):
        """Validate export configuration"""
        if not self.source_report:
            frappe.throw(_("Source report is required"))

        # Validate source report exists
        if not frappe.db.exists("Custom Report Builder", self.source_report):
            frappe.throw(_("Source report '{0}' not found").format(self.source_report))

        # Validate output format
        if not self.output_format:
            frappe.throw(_("Output format is required"))

    def validate_arabic_settings(self):
        """Validate Arabic localization settings"""
        if self.rtl_layout:
            if not self.arabic_fonts:
                self.arabic_fonts = "Noto Sans Arabic"

            if not self.date_format_arabic:
                self.date_format_arabic = "DD/MM/YYYY"

            if not self.currency_format_arabic:
                self.currency_format_arabic = "OMR 123.456"

    def validate_mobile_settings(self):
        """Validate mobile optimization settings"""
        if self.mobile_friendly and self.output_format == "PDF":
            # Ensure mobile-friendly PDF settings
            if not self.page_orientation:
                self.page_orientation = "Portrait"

            if not self.paper_size:
                self.paper_size = "A4"

    def set_default_values(self):
        """Set default values for fields"""
        if not self.export_date:
            self.export_date = datetime.now()

        if not self.retention_days:
            self.retention_days = 30

        if not self.security_level:
            self.security_level = "Internal"

    @frappe.whitelist()
    def generate_export(self) -> Dict[str, Any]:
        """Generate export file based on configuration"""
        try:
            self.status = "Generating"
            self.save()

            start_time = time.time()

            # Get source report data
            source_data = self.get_source_report_data()

            # Generate export based on format
            export_result = self.create_export_file(source_data)

            # Calculate duration and metadata
            end_time = time.time()
            duration = end_time - start_time

            # Update document with results
            self.export_duration = f"{duration:.2f} seconds"
            self.export_size_mb = export_result.get("file_size_mb", 0)
            self.export_rows = export_result.get("row_count", 0)
            self.generated_file_path = export_result.get("file_path", "")
            self.download_url = export_result.get("download_url", "")
            self.file_checksum = export_result.get("checksum", "")
            self.status = "Completed"
            self.success_message = _("Export generated successfully")

            self.save()

            # Handle delivery options
            if self.email_delivery:
                self.send_email_delivery()

            if self.cloud_storage:
                self.upload_to_cloud_storage()

            return {
                "success": True,
                "message": _("Export generated successfully"),
                "download_url": self.download_url,
                "file_size": self.file_size,
                "export_id": self.name,
            }

        except Exception as e:
            self.status = "Failed"
            self.error_log = str(e)
            self.save()
            frappe.log_error(f"Export generation failed: {e}", "Report Export Utility")
            return {"success": False, "message": _("Export generation failed: {0}").format(str(e))}

    def get_source_report_data(self) -> Dict[str, Any]:
        """Get data from source report"""
        report = frappe.get_doc("Custom Report Builder", self.source_report)

        # Execute report query
        if report.sql_query:
            data = frappe.db.sql(report.sql_query, as_dict=True)
        else:
            # Use report configuration to build query
            data = self.build_report_data(report)

        return {
            "data": data,
            "report": report,
            "metadata": {
                "generated_on": datetime.now(),
                "row_count": len(data),
                "columns": list(data[0].keys()) if data else [],
            },
        }

    def build_report_data(self, report) -> List[Dict[str, Any]]:
        """Build report data from configuration"""
        filters = {}

        if hasattr(report, "report_filters") and report.report_filters:
            try:
                filters = json.loads(report.report_filters)
            except json.JSONDecodeError:
                pass

        # Get data from source DocType
        if hasattr(report, "source_doctype") and report.source_doctype:
            fields = getattr(report, "selected_fields", "*")
            if isinstance(fields, str) and fields != "*":
                try:
                    fields = json.loads(fields)
                except json.JSONDecodeError:
                    fields = ["*"]

            data = frappe.get_list(
                report.source_doctype,
                filters=filters,
                fields=fields,
                limit=getattr(report, "row_limit", 1000),
            )

            return data

        return []

    def create_export_file(self, source_data: Dict[str, Any]) -> Dict[str, Any]:
        """Create export file based on output format"""
        format_handlers = {
            "PDF": self.generate_pdf_export,
            "Excel (XLSX)": self.generate_excel_export,
            "CSV": self.generate_csv_export,
            "HTML": self.generate_html_export,
            "JSON": self.generate_json_export,
            "XML": self.generate_xml_export,
            "PowerPoint": self.generate_powerpoint_export,
        }

        handler = format_handlers.get(self.output_format)
        if not handler:
            frappe.throw(_("Unsupported export format: {0}").format(self.output_format))

        return handler(source_data)

    def generate_pdf_export(self, source_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate PDF export with Arabic RTL support"""
        html_content = self.build_html_content(source_data)

        # Configure PDF options
        options = self.get_pdf_options()

        # Generate PDF
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_file:
            try:
                pdfkit.from_string(html_content, tmp_file.name, options=options)

                # Calculate file size
                file_size = os.path.getsize(tmp_file.name)
                file_size_mb = file_size / (1024 * 1024)

                # Generate checksum
                checksum = self.calculate_file_checksum(tmp_file.name)

                # Move to final location
                final_path = self.save_export_file(tmp_file.name, "pdf")

                return {
                    "file_path": final_path,
                    "file_size_mb": file_size_mb,
                    "checksum": checksum,
                    "row_count": len(source_data["data"]),
                    "download_url": self.generate_download_url(final_path),
                }

            except Exception as e:
                if os.path.exists(tmp_file.name):
                    os.unlink(tmp_file.name)
                raise e

    def build_html_content(self, source_data: Dict[str, Any]) -> str:
        """Build HTML content for PDF export with Arabic support"""
        report = source_data["report"]
        data = source_data["data"]

        # Build CSS styles
        css_styles = self.build_css_styles()

        # Build header
        html = f"""
        <!DOCTYPE html>
        <html dir="{self.get_text_direction()}">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>{self.get_report_title(report)}</title>
            <style>{css_styles}</style>
        </head>
        <body>
        """

        # Add header
        html += self.build_report_header(report)

        # Add filters if requested
        if self.include_filters:
            html += self.build_filters_section(report)

        # Add data table
        html += self.build_data_table(data)

        # Add summary if requested
        if self.include_summary:
            html += self.build_summary_section(data)

        # Add charts if requested
        if self.include_charts and hasattr(report, "chart_config"):
            html += self.build_charts_section(report, data)

        # Add footer
        html += self.build_report_footer()

        html += """
        </body>
        </html>
        """

        return html

    def build_css_styles(self) -> str:
        """Build CSS styles for export"""
        css = """
        * { margin: 0; padding: 0; box-sizing: border-box; }
        
        body {
            font-family: 'Noto Sans Arabic', Tahoma, Arial, sans-serif;
            font-size: 12px;
            line-height: 1.4;
            color: #333;
        }
        """

        # RTL styles
        if self.rtl_layout:
            css += """
            body { direction: rtl; text-align: right; }
            .table { direction: rtl; }
            .table th, .table td { text-align: right; }
            .number { direction: ltr; text-align: left; }
            """

        # Mobile optimization
        if self.mobile_friendly:
            css += """
            @media screen and (max-width: 768px) {
                body { font-size: 14px; }
                .table { font-size: 10px; }
                .responsive-table { 
                    overflow-x: auto; 
                    white-space: nowrap; 
                }
            }
            """

        # Custom fonts for Arabic
        if self.rtl_layout and self.arabic_fonts:
            font_family = self.arabic_fonts.replace(" ", "\\ ")
            css += f"""
            .arabic-text {{ font-family: '{font_family}', Tahoma, Arial; }}
            """

        # Table styles
        css += """
        .table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }
        
        .table th, .table td {
            border: 1px solid #ddd;
            padding: 8px;
            vertical-align: top;
        }
        
        .table th {
            background-color: #f5f5f5;
            font-weight: bold;
        }
        
        .header {
            text-align: center;
            margin-bottom: 30px;
            border-bottom: 2px solid #007bff;
            padding-bottom: 20px;
        }
        
        .footer {
            margin-top: 30px;
            border-top: 1px solid #ddd;
            padding-top: 10px;
            font-size: 10px;
            color: #666;
        }
        """

        return css

    def get_text_direction(self) -> str:
        """Get text direction for HTML"""
        return "rtl" if self.rtl_layout else "ltr"

    def get_report_title(self, report) -> str:
        """Get report title with Arabic support"""
        if self.rtl_layout and hasattr(report, "report_name_ar") and report.report_name_ar:
            return report.report_name_ar
        return getattr(report, "report_name", self.export_name)

    def build_report_header(self, report) -> str:
        """Build report header section"""
        title = self.get_report_title(report)
        current_date = self.format_date_for_display(datetime.now())

        header = f"""
        <div class="header">
            <h1>{title}</h1>
            <p>{_("Generated on")}: {current_date}</p>
            <p>{_("Export Format")}: {self.output_format}</p>
        </div>
        """

        return header

    def build_data_table(self, data: List[Dict[str, Any]]) -> str:
        """Build data table HTML"""
        if not data:
            return f"<p>{_('No data available')}</p>"

        # Get column headers
        columns = list(data[0].keys())

        table_class = "table"
        if self.mobile_friendly:
            table_class += " responsive-table"

        html = f'<div class="{table_class}"><table class="table">'

        # Table header
        html += "<thead><tr>"
        for col in columns:
            display_name = self.get_column_display_name(col)
            html += f"<th>{display_name}</th>"
        html += "</tr></thead>"

        # Table body
        html += "<tbody>"
        for row in data:
            html += "<tr>"
            for col in columns:
                value = row.get(col, "")
                formatted_value = self.format_cell_value(value, col)
                html += f"<td>{formatted_value}</td>"
            html += "</tr>"
        html += "</tbody>"

        html += "</table></div>"

        return html

    def format_cell_value(self, value: Any, column_name: str) -> str:
        """Format cell value with Arabic localization"""
        if value is None:
            return ""

        # Handle dates
        if isinstance(value, (datetime, type(datetime.now().date()))):
            return self.format_date_for_display(value)

        # Handle numbers
        if isinstance(value, (int, float)):
            if "currency" in column_name.lower() or "amount" in column_name.lower():
                return self.format_currency_for_display(value)
            else:
                return self.format_number_for_display(value)

        # Handle strings
        return str(value)

    def format_date_for_display(self, date_value) -> str:
        """Format date for display with Arabic support"""
        if not date_value:
            return ""

        if self.rtl_layout and self.date_format_arabic == "Hijri Calendar":
            # Implementation for Hijri calendar would go here
            # For now, use Gregorian with Arabic format
            pass

        # Default formatting
        if isinstance(date_value, str):
            return date_value

        if self.rtl_layout:
            formatted = date_value.strftime("%d/%m/%Y")
            if self.arabic_numbers:
                formatted = self.convert_to_arabic_numerals(formatted)
            return formatted

        return date_value.strftime("%Y-%m-%d")

    def format_currency_for_display(self, amount: float) -> str:
        """Format currency for display with Arabic support"""
        if self.rtl_layout and self.currency_format_arabic:
            if self.currency_format_arabic == "ر.ع. ١٢٣.٤٥٦":
                formatted = f"ر.ع. {amount:,.3f}"
                if self.arabic_numbers:
                    formatted = self.convert_to_arabic_numerals(formatted)
                return formatted
            elif self.currency_format_arabic == "١٢٣.٤٥٦ ر.ع.":
                formatted = f"{amount:,.3f} ر.ع."
                if self.arabic_numbers:
                    formatted = self.convert_to_arabic_numerals(formatted)
                return formatted

        # Default format
        return f"OMR {amount:,.3f}"

    def format_number_for_display(self, number: float) -> str:
        """Format number for display with Arabic support"""
        formatted = f"{number:,.2f}"

        if self.rtl_layout and self.arabic_numbers:
            formatted = self.convert_to_arabic_numerals(formatted)

        return formatted

    def convert_to_arabic_numerals(self, text: str) -> str:
        """Convert Western numerals to Arabic-Indic numerals"""
        arabic_numerals = {
            "0": "٠",
            "1": "١",
            "2": "٢",
            "3": "٣",
            "4": "٤",
            "5": "٥",
            "6": "٦",
            "7": "٧",
            "8": "٨",
            "9": "٩",
        }

        for western, arabic in arabic_numerals.items():
            text = text.replace(western, arabic)

        return text

    def get_column_display_name(self, column_name: str) -> str:
        """Get display name for column with translation"""
        # Try to get translated name
        translated = _(column_name.replace("_", " ").title())
        return translated

    def get_pdf_options(self) -> Dict[str, Any]:
        """Get PDF generation options"""
        options = {
            "page-size": self.paper_size or "A4",
            "orientation": self.page_orientation or "Portrait",
            "encoding": "UTF-8",
            "no-outline": None,
            "enable-local-file-access": None,
        }

        # Mobile optimization
        if self.mobile_friendly:
            options.update({"minimum-font-size": "12", "zoom": "0.8"})

        # Custom margins
        if self.margin_settings:
            try:
                margins = self.margin_settings.split()
                if len(margins) >= 2:
                    options["margin-top"] = margins[0]
                    options["margin-right"] = margins[1]
                    if len(margins) >= 3:
                        options["margin-bottom"] = margins[2]
                    if len(margins) >= 4:
                        options["margin-left"] = margins[3]
            except:
                pass

        return options

    def generate_excel_export(self, source_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate Excel export with Arabic RTL support"""
        data = source_data["data"]
        report = source_data["report"]

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = self.get_report_title(report)[:31]  # Excel sheet name limit

        if not data:
            ws["A1"] = _("No data available")
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
                wb.save(tmp_file.name)
                return self.finalize_export_file(tmp_file.name, "xlsx", 0)

        # Set RTL if needed
        if self.rtl_layout:
            ws.sheet_view.rightToLeft = True

        # Add headers
        columns = list(data[0].keys())
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = self.get_column_display_name(col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right" if self.rtl_layout else "left")

        # Add data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(col_name, "")

                # Format value for Excel
                if isinstance(value, (datetime, type(datetime.now().date()))):
                    cell.value = value
                    cell.number_format = "DD/MM/YYYY" if self.rtl_layout else "YYYY-MM-DD"
                elif isinstance(value, (int, float)):
                    cell.value = value
                    if "currency" in col_name.lower():
                        cell.number_format = "#,##0.000"
                else:
                    cell.value = str(value)

                cell.alignment = Alignment(horizontal="right" if self.rtl_layout else "left")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            wb.save(tmp_file.name)
            return self.finalize_export_file(tmp_file.name, "xlsx", len(data))

    def generate_csv_export(self, source_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate CSV export with Arabic support"""
        data = source_data["data"]

        if not data:
            # Create empty CSV
            with tempfile.NamedTemporaryFile(
                mode="w", suffix=".csv", delete=False, encoding="utf-8-sig"
            ) as tmp_file:
                tmp_file.write(_("No data available"))
                return self.finalize_export_file(tmp_file.name, "csv", 0)

        # Create CSV
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False, encoding="utf-8-sig"
        ) as tmp_file:
            columns = list(data[0].keys())

            # Write headers
            headers = [self.get_column_display_name(col) for col in columns]
            writer = csv.writer(tmp_file)
            writer.writerow(headers)

            # Write data
            for row in data:
                formatted_row = []
                for col in columns:
                    value = row.get(col, "")
                    if isinstance(value, (datetime, type(datetime.now().date()))):
                        value = self.format_date_for_display(value)
                    elif isinstance(value, (int, float)):
                        value = self.format_number_for_display(value)
                    formatted_row.append(str(value))
                writer.writerow(formatted_row)

            return self.finalize_export_file(tmp_file.name, "csv", len(data))

    def generate_html_export(self, source_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate HTML export"""
        html_content = self.build_html_content(source_data)

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".html", delete=False, encoding="utf-8"
        ) as tmp_file:
            tmp_file.write(html_content)
            return self.finalize_export_file(tmp_file.name, "html", len(source_data["data"]))

    def generate_json_export(self, source_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate JSON export"""
        export_data = {
            "report_name": self.get_report_title(source_data["report"]),
            "generated_on": datetime.now().isoformat(),
            "export_config": {
                "format": self.output_format,
                "rtl_layout": self.rtl_layout,
                "mobile_friendly": self.mobile_friendly,
            },
            "data": source_data["data"],
            "metadata": source_data["metadata"],
        }

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".json", delete=False, encoding="utf-8"
        ) as tmp_file:
            json.dump(export_data, tmp_file, ensure_ascii=False, indent=2, default=str)
            return self.finalize_export_file(tmp_file.name, "json", len(source_data["data"]))

    def generate_xml_export(self, source_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate XML export"""
        # Simple XML generation - for more complex XML, consider using lxml
        xml_content = '<?xml version="1.0" encoding="UTF-8"?>\n'
        xml_content += f'<report name="{self.get_report_title(source_data["report"])}">\n'
        xml_content += f'  <metadata generated_on="{datetime.now().isoformat()}" />\n'
        xml_content += "  <data>\n"

        for row in source_data["data"]:
            xml_content += "    <row>\n"
            for key, value in row.items():
                safe_key = key.replace(" ", "_").replace("-", "_")
                xml_content += f"      <{safe_key}>{str(value)}</{safe_key}>\n"
            xml_content += "    </row>\n"

        xml_content += "  </data>\n"
        xml_content += "</report>\n"

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".xml", delete=False, encoding="utf-8"
        ) as tmp_file:
            tmp_file.write(xml_content)
            return self.finalize_export_file(tmp_file.name, "xml", len(source_data["data"]))

    def generate_powerpoint_export(self, source_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate PowerPoint export - placeholder for future implementation"""
        # This would require python-pptx library
        frappe.throw(_("PowerPoint export not yet implemented"))

    def finalize_export_file(
        self, temp_path: str, extension: str, row_count: int
    ) -> Dict[str, Any]:
        """Finalize export file and move to final location"""
        try:
            # Calculate file size
            file_size = os.path.getsize(temp_path)
            file_size_mb = file_size / (1024 * 1024)

            # Generate checksum
            checksum = self.calculate_file_checksum(temp_path)

            # Move to final location
            final_path = self.save_export_file(temp_path, extension)

            return {
                "file_path": final_path,
                "file_size_mb": file_size_mb,
                "checksum": checksum,
                "row_count": row_count,
                "download_url": self.generate_download_url(final_path),
            }

        finally:
            # Clean up temp file
            if os.path.exists(temp_path):
                os.unlink(temp_path)

    def calculate_file_checksum(self, file_path: str) -> str:
        """Calculate MD5 checksum of file"""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()

    def save_export_file(self, temp_path: str, extension: str) -> str:
        """Save export file to final location"""
        # Create exports directory if it doesn't exist
        exports_dir = os.path.join(frappe.utils.get_site_path(), "private", "files", "exports")
        if not os.path.exists(exports_dir):
            os.makedirs(exports_dir)

        # Generate unique filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.name}_{timestamp}.{extension}"
        final_path = os.path.join(exports_dir, filename)

        # Move file
        import shutil

        shutil.move(temp_path, final_path)

        # Update file size
        file_size = os.path.getsize(final_path)
        self.file_size = self.format_file_size(file_size)

        return final_path

    def format_file_size(self, size_bytes: int) -> str:
        """Format file size for display"""
        if size_bytes < 1024:
            return f"{size_bytes} B"
        elif size_bytes < 1024 * 1024:
            return f"{size_bytes / 1024:.1f} KB"
        elif size_bytes < 1024 * 1024 * 1024:
            return f"{size_bytes / (1024 * 1024):.1f} MB"
        else:
            return f"{size_bytes / (1024 * 1024 * 1024):.1f} GB"

    def generate_download_url(self, file_path: str) -> str:
        """Generate download URL for file"""
        # This would generate a secure download URL
        # For now, return a placeholder
        return f"/api/method/universal_workshop.analytics_reporting.doctype.report_export_utility.report_export_utility.download_export?export_id={self.name}"

    def send_email_delivery(self):
        """Send export file via email"""
        if not self.email_recipients or not self.generated_file_path:
            return

        recipients = [email.strip() for email in self.email_recipients.split(",")]

        subject = _("Report Export: {0}").format(self.export_name)

        message = f"""
        <p>{_("Dear User")},</p>
        <p>{_("Your requested report export is ready for download.")} </p>
        <p><strong>{_("Report Name")}:</strong> {self.export_name}</p>
        <p><strong>{_("Export Format")}:</strong> {self.output_format}</p>
        <p><strong>{_("Generated On")}:</strong> {self.export_date}</p>
        <p><strong>{_("File Size")}:</strong> {self.file_size}</p>
        <p>{_("Please find the export file attached.")}</p>
        <p>{_("Best regards")},<br/>{_("Workshop Management System")}</p>
        """

        try:
            frappe.sendmail(
                recipients=recipients,
                subject=subject,
                message=message,
                attachments=[
                    {
                        "fname": os.path.basename(self.generated_file_path),
                        "fcontent": open(self.generated_file_path, "rb").read(),
                    }
                ],
                reference_doctype="Report Export Utility",
                reference_name=self.name,
            )
        except Exception as e:
            frappe.log_error(f"Failed to send export email: {e}", "Report Export Utility")

    def upload_to_cloud_storage(self):
        """Upload export file to cloud storage"""
        # Placeholder for cloud storage implementation
        # This would integrate with AWS S3, Google Cloud Storage, etc.
        pass


# Utility functions


@frappe.whitelist()
def download_export(export_id: str):
    """Download export file"""
    export_doc = frappe.get_doc("Report Export Utility", export_id)

    # Check permissions
    if not frappe.has_permission("Report Export Utility", "read", export_doc):
        frappe.throw(_("No permission to download this export"))

    # Check if file exists
    if not export_doc.generated_file_path or not os.path.exists(export_doc.generated_file_path):
        frappe.throw(_("Export file not found"))

    # Check if expired
    if export_doc.is_expired or (
        export_doc.expiry_date and export_doc.expiry_date < datetime.now()
    ):
        frappe.throw(_("Export file has expired"))

    # Increment download count
    export_doc.download_count = (export_doc.download_count or 0) + 1
    export_doc.save(ignore_permissions=True)

    # Return file for download
    from frappe.utils.file_manager import get_file

    filename = os.path.basename(export_doc.generated_file_path)

    with open(export_doc.generated_file_path, "rb") as f:
        frappe.local.response.filename = filename
        frappe.local.response.filecontent = f.read()
        frappe.local.response.type = "download"


@frappe.whitelist()
def create_quick_export(source_report: str, output_format: str = "PDF", **kwargs) -> Dict[str, Any]:
    """Create a quick export without saving configuration"""
    export_doc = frappe.new_doc("Report Export Utility")
    export_doc.export_name = f"Quick Export - {datetime.now().strftime('%Y%m%d_%H%M%S')}"
    export_doc.source_report = source_report
    export_doc.output_format = output_format

    # Apply any additional settings
    for key, value in kwargs.items():
        if hasattr(export_doc, key):
            setattr(export_doc, key, value)

    export_doc.insert(ignore_permissions=True)

    return export_doc.generate_export()


@frappe.whitelist()
def get_export_formats() -> List[str]:
    """Get available export formats"""
    return ["PDF", "Excel (XLSX)", "CSV", "HTML", "JSON", "XML"]


@frappe.whitelist()
def cleanup_expired_exports():
    """Cleanup expired export files"""
    expired_exports = frappe.get_list(
        "Report Export Utility",
        filters={"expiry_date": ["<", datetime.now()], "is_expired": 0},
        fields=["name", "generated_file_path"],
    )

    for export in expired_exports:
        try:
            # Mark as expired
            doc = frappe.get_doc("Report Export Utility", export.name)
            doc.is_expired = 1
            doc.save(ignore_permissions=True)

            # Delete file
            if export.generated_file_path and os.path.exists(export.generated_file_path):
                os.unlink(export.generated_file_path)

        except Exception as e:
            frappe.log_error(
                f"Failed to cleanup export {export.name}: {e}", "Report Export Utility"
            )

    return len(expired_exports)
