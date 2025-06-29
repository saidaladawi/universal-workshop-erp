"""
Cost Analysis System for Procurement and Supplier Management
Universal Workshop ERP - Advanced Analytics with Arabic Localization
"""

import frappe
from frappe import _
from frappe.utils import flt, cint, getdate, add_days, add_months, get_first_day, get_last_day
from frappe.utils.data import today, get_datetime
import json
from datetime import datetime, timedelta
import pandas as pd


class ProcurementCostAnalyzer:
    """Advanced cost analysis system for procurement analytics"""

    def __init__(self):
        self.currency = frappe.defaults.get_global_default("currency") or "OMR"
        self.company = frappe.defaults.get_global_default("company")

    @frappe.whitelist()
    def get_supplier_spend_analysis(self, from_date=None, to_date=None, supplier=None):
        """Get comprehensive supplier spend analysis"""

        if not from_date:
            from_date = add_months(today(), -12)
        if not to_date:
            to_date = today()

        # Base query for supplier spend
        conditions = {"docstatus": 1, "posting_date": ["between", [from_date, to_date]]}

        if supplier:
            conditions["supplier"] = supplier

        # Purchase Orders data
        po_data = frappe.db.sql(
            """
            SELECT 
                supplier,
                supplier_name,
                COUNT(*) as total_orders,
                SUM(grand_total) as total_spend,
                AVG(grand_total) as avg_order_value,
                MIN(grand_total) as min_order_value,
                MAX(grand_total) as max_order_value,
                SUM(total_qty) as total_qty,
                COUNT(DISTINCT MONTH(posting_date)) as active_months
            FROM `tabPurchase Order`
            WHERE docstatus = 1 
                AND posting_date BETWEEN %(from_date)s AND %(to_date)s
                {supplier_condition}
            GROUP BY supplier, supplier_name
            ORDER BY total_spend DESC
        """.format(
                supplier_condition="AND supplier = %(supplier)s" if supplier else ""
            ),
            {"from_date": from_date, "to_date": to_date, "supplier": supplier},
            as_dict=True,
        )

        # Purchase Receipts data for quality insights
        pr_data = frappe.db.sql(
            """
            SELECT 
                supplier,
                COUNT(*) as total_receipts,
                AVG(CASE WHEN quality_inspection_status = 'Completed' THEN 1 ELSE 0 END) as quality_pass_rate,
                COUNT(CASE WHEN quality_inspection_status = 'Failed' THEN 1 END) as failed_inspections
            FROM `tabPurchase Receipt`
            WHERE docstatus = 1 
                AND posting_date BETWEEN %(from_date)s AND %(to_date)s
                {supplier_condition}
            GROUP BY supplier
        """.format(
                supplier_condition="AND supplier = %(supplier)s" if supplier else ""
            ),
            {"from_date": from_date, "to_date": to_date, "supplier": supplier},
            as_dict=True,
        )

        # Merge quality data with spend data
        quality_dict = {item["supplier"]: item for item in pr_data}

        for spend_item in po_data:
            quality_info = quality_dict.get(spend_item["supplier"], {})
            spend_item.update(
                {
                    "quality_pass_rate": flt(quality_info.get("quality_pass_rate", 0)) * 100,
                    "failed_inspections": quality_info.get("failed_inspections", 0),
                    "total_receipts": quality_info.get("total_receipts", 0),
                }
            )

        # Calculate spend percentages
        total_spend = sum([item["total_spend"] for item in po_data])

        for item in po_data:
            item["spend_percentage"] = (
                flt(item["total_spend"] / total_spend * 100, 2) if total_spend > 0 else 0
            )
            item["currency"] = self.currency

        return {
            "supplier_spend_data": po_data,
            "summary": {
                "total_suppliers": len(po_data),
                "total_spend": total_spend,
                "total_orders": sum([item["total_orders"] for item in po_data]),
                "avg_spend_per_supplier": flt(total_spend / len(po_data), 2) if po_data else 0,
                "currency": self.currency,
                "period": f"{from_date} to {to_date}",
            },
        }

    @frappe.whitelist()
    def get_price_trend_analysis(
        self, item_code=None, item_group=None, from_date=None, to_date=None
    ):
        """Get price trend analysis for items"""

        if not from_date:
            from_date = add_months(today(), -12)
        if not to_date:
            to_date = today()

        conditions = ["po.docstatus = 1", "po.posting_date BETWEEN %(from_date)s AND %(to_date)s"]

        if item_code:
            conditions.append("poi.item_code = %(item_code)s")
        if item_group:
            conditions.append("i.item_group = %(item_group)s")

        conditions_str = " AND ".join(conditions)

        price_data = frappe.db.sql(
            f"""
            SELECT 
                poi.item_code,
                poi.item_name,
                i.item_group,
                po.posting_date,
                po.supplier,
                po.supplier_name,
                poi.rate,
                poi.qty,
                poi.amount,
                (poi.amount / poi.qty) as unit_cost,
                MONTH(po.posting_date) as month,
                YEAR(po.posting_date) as year
            FROM `tabPurchase Order Item` poi
            JOIN `tabPurchase Order` po ON poi.parent = po.name
            JOIN `tabItem` i ON poi.item_code = i.item_code
            WHERE {conditions_str}
            ORDER BY poi.item_code, po.posting_date
        """,
            {
                "from_date": from_date,
                "to_date": to_date,
                "item_code": item_code,
                "item_group": item_group,
            },
            as_dict=True,
        )

        # Process price trends
        price_trends = {}

        for record in price_data:
            item_key = record["item_code"]

            if item_key not in price_trends:
                price_trends[item_key] = {
                    "item_code": record["item_code"],
                    "item_name": record["item_name"],
                    "item_group": record["item_group"],
                    "price_history": [],
                    "suppliers": set(),
                    "price_stats": {},
                }

            price_trends[item_key]["price_history"].append(
                {
                    "date": record["posting_date"],
                    "rate": record["rate"],
                    "supplier": record["supplier"],
                    "supplier_name": record["supplier_name"],
                    "qty": record["qty"],
                    "amount": record["amount"],
                }
            )

            price_trends[item_key]["suppliers"].add(record["supplier_name"])

        # Calculate price statistics
        for item_key, trend_data in price_trends.items():
            prices = [item["rate"] for item in trend_data["price_history"]]

            if prices:
                trend_data["price_stats"] = {
                    "min_price": min(prices),
                    "max_price": max(prices),
                    "avg_price": sum(prices) / len(prices),
                    "latest_price": prices[-1] if prices else 0,
                    "price_variance": (
                        ((max(prices) - min(prices)) / min(prices) * 100) if min(prices) > 0 else 0
                    ),
                    "total_suppliers": len(trend_data["suppliers"]),
                    "currency": self.currency,
                }

            # Convert suppliers set to list for JSON serialization
            trend_data["suppliers"] = list(trend_data["suppliers"])

        return {
            "price_trends": list(price_trends.values()),
            "summary": {
                "total_items": len(price_trends),
                "period": f"{from_date} to {to_date}",
                "currency": self.currency,
            },
        }

    @frappe.whitelist()
    def get_cost_breakdown_analysis(self, from_date=None, to_date=None):
        """Get comprehensive cost breakdown analysis"""

        if not from_date:
            from_date = add_months(today(), -12)
        if not to_date:
            to_date = today()

        # Cost breakdown by item group
        item_group_data = frappe.db.sql(
            """
            SELECT 
                i.item_group,
                COUNT(DISTINCT poi.item_code) as total_items,
                COUNT(*) as total_transactions,
                SUM(poi.amount) as total_spend,
                AVG(poi.rate) as avg_rate,
                SUM(poi.qty) as total_qty
            FROM `tabPurchase Order Item` poi
            JOIN `tabPurchase Order` po ON poi.parent = po.name
            JOIN `tabItem` i ON poi.item_code = i.item_code
            WHERE po.docstatus = 1 
                AND po.posting_date BETWEEN %(from_date)s AND %(to_date)s
            GROUP BY i.item_group
            ORDER BY total_spend DESC
        """,
            {"from_date": from_date, "to_date": to_date},
            as_dict=True,
        )

        # Monthly spend trend
        monthly_data = frappe.db.sql(
            """
            SELECT 
                YEAR(posting_date) as year,
                MONTH(posting_date) as month,
                DATE_FORMAT(posting_date, '%%Y-%%m') as period,
                COUNT(*) as total_orders,
                SUM(grand_total) as total_spend,
                COUNT(DISTINCT supplier) as unique_suppliers
            FROM `tabPurchase Order`
            WHERE docstatus = 1 
                AND posting_date BETWEEN %(from_date)s AND %(to_date)s
            GROUP BY year, month
            ORDER BY year, month
        """,
            {"from_date": from_date, "to_date": to_date},
            as_dict=True,
        )

        # Top performing suppliers (by spend and quality)
        top_suppliers = frappe.db.sql(
            """
            SELECT 
                po.supplier,
                po.supplier_name,
                COUNT(*) as total_orders,
                SUM(po.grand_total) as total_spend,
                AVG(po.grand_total) as avg_order_value,
                COUNT(DISTINCT DATE(po.posting_date)) as active_days
            FROM `tabPurchase Order` po
            WHERE po.docstatus = 1 
                AND po.posting_date BETWEEN %(from_date)s AND %(to_date)s
            GROUP BY po.supplier, po.supplier_name
            ORDER BY total_spend DESC
            LIMIT 10
        """,
            {"from_date": from_date, "to_date": to_date},
            as_dict=True,
        )

        # Calculate totals and percentages
        total_spend = sum([item["total_spend"] for item in item_group_data])

        for item in item_group_data:
            item["spend_percentage"] = (
                flt(item["total_spend"] / total_spend * 100, 2) if total_spend > 0 else 0
            )
            item["currency"] = self.currency

        for supplier in top_suppliers:
            supplier["currency"] = self.currency

        return {
            "cost_by_item_group": item_group_data,
            "monthly_trend": monthly_data,
            "top_suppliers": top_suppliers,
            "summary": {
                "total_spend": total_spend,
                "total_item_groups": len(item_group_data),
                "period": f"{from_date} to {to_date}",
                "currency": self.currency,
            },
        }

    @frappe.whitelist()
    def get_supplier_performance_scorecard(self, supplier=None, from_date=None, to_date=None):
        """Get comprehensive supplier performance scorecard"""

        if not from_date:
            from_date = add_months(today(), -12)
        if not to_date:
            to_date = today()

        conditions = ["po.docstatus = 1", "po.posting_date BETWEEN %(from_date)s AND %(to_date)s"]

        if supplier:
            conditions.append("po.supplier = %(supplier)s")

        conditions_str = " AND ".join(conditions)

        performance_data = frappe.db.sql(
            f"""
            SELECT 
                po.supplier,
                po.supplier_name,
                COUNT(*) as total_orders,
                SUM(po.grand_total) as total_spend,
                AVG(po.grand_total) as avg_order_value,
                AVG(DATEDIFF(po.posting_date, po.transaction_date)) as avg_delivery_days,
                COUNT(CASE WHEN po.status = 'Completed' THEN 1 END) as completed_orders,
                COUNT(CASE WHEN po.status = 'Cancelled' THEN 1 END) as cancelled_orders
            FROM `tabPurchase Order` po
            WHERE {conditions_str}
            GROUP BY po.supplier, po.supplier_name
            ORDER BY total_spend DESC
        """,
            {"from_date": from_date, "to_date": to_date, "supplier": supplier},
            as_dict=True,
        )

        # Add quality metrics from Purchase Receipts
        for supplier_data in performance_data:
            quality_metrics = frappe.db.sql(
                """
                SELECT 
                    COUNT(*) as total_receipts,
                    COUNT(CASE WHEN quality_inspection_status = 'Completed' THEN 1 END) as passed_inspections,
                    COUNT(CASE WHEN quality_inspection_status = 'Failed' THEN 1 END) as failed_inspections,
                    AVG(CASE WHEN quality_inspection_status IN ('Completed', 'Failed') THEN 1 ELSE 0 END) as inspection_rate
                FROM `tabPurchase Receipt`
                WHERE supplier = %(supplier)s 
                    AND docstatus = 1 
                    AND posting_date BETWEEN %(from_date)s AND %(to_date)s
            """,
                {"supplier": supplier_data["supplier"], "from_date": from_date, "to_date": to_date},
                as_dict=True,
            )[0]

            # Calculate performance scores
            delivery_score = max(
                0, 100 - (supplier_data["avg_delivery_days"] or 0) * 2
            )  # Penalty for late delivery
            completion_rate = (
                (supplier_data["completed_orders"] / supplier_data["total_orders"] * 100)
                if supplier_data["total_orders"] > 0
                else 0
            )
            quality_rate = (
                (quality_metrics["passed_inspections"] / quality_metrics["total_receipts"] * 100)
                if quality_metrics["total_receipts"] > 0
                else 0
            )

            supplier_data.update(
                {
                    "quality_metrics": quality_metrics,
                    "delivery_score": flt(delivery_score, 2),
                    "completion_rate": flt(completion_rate, 2),
                    "quality_rate": flt(quality_rate, 2),
                    "overall_score": flt((delivery_score + completion_rate + quality_rate) / 3, 2),
                    "currency": self.currency,
                }
            )

        return {
            "supplier_scorecards": performance_data,
            "summary": {
                "total_suppliers": len(performance_data),
                "period": f"{from_date} to {to_date}",
                "currency": self.currency,
            },
        }

    @frappe.whitelist()
    def get_procurement_kpis(self, from_date=None, to_date=None):
        """Get key performance indicators for procurement"""

        if not from_date:
            from_date = add_months(today(), -12)
        if not to_date:
            to_date = today()

        # Previous period for comparison
        prev_from_date = add_months(from_date, -12)
        prev_to_date = add_months(to_date, -12)

        current_kpis = self._calculate_period_kpis(from_date, to_date)
        previous_kpis = self._calculate_period_kpis(prev_from_date, prev_to_date)

        # Calculate growth rates
        for key in current_kpis:
            if key in previous_kpis and previous_kpis[key] > 0:
                current_kpis[f"{key}_growth"] = flt(
                    (current_kpis[key] - previous_kpis[key]) / previous_kpis[key] * 100, 2
                )
            else:
                current_kpis[f"{key}_growth"] = 0

        return {
            "current_period": current_kpis,
            "previous_period": previous_kpis,
            "period": f"{from_date} to {to_date}",
            "comparison_period": f"{prev_from_date} to {prev_to_date}",
            "currency": self.currency,
        }

    def _calculate_period_kpis(self, from_date, to_date):
        """Calculate KPIs for a specific period"""

        kpis = frappe.db.sql(
            """
            SELECT 
                COUNT(*) as total_orders,
                SUM(grand_total) as total_spend,
                AVG(grand_total) as avg_order_value,
                COUNT(DISTINCT supplier) as unique_suppliers,
                COUNT(CASE WHEN status = 'Completed' THEN 1 END) as completed_orders,
                COUNT(CASE WHEN status = 'Cancelled' THEN 1 END) as cancelled_orders,
                AVG(DATEDIFF(posting_date, transaction_date)) as avg_processing_days
            FROM `tabPurchase Order`
            WHERE docstatus = 1 
                AND posting_date BETWEEN %(from_date)s AND %(to_date)s
        """,
            {"from_date": from_date, "to_date": to_date},
            as_dict=True,
        )[0]

        # Quality KPIs
        quality_kpis = frappe.db.sql(
            """
            SELECT 
                COUNT(*) as total_receipts,
                COUNT(CASE WHEN quality_inspection_status = 'Completed' THEN 1 END) as passed_inspections,
                COUNT(CASE WHEN quality_inspection_status = 'Failed' THEN 1 END) as failed_inspections,
                AVG(CASE WHEN quality_inspection_status = 'Completed' THEN 1 ELSE 0 END) as quality_pass_rate
            FROM `tabPurchase Receipt`
            WHERE docstatus = 1 
                AND posting_date BETWEEN %(from_date)s AND %(to_date)s
        """,
            {"from_date": from_date, "to_date": to_date},
            as_dict=True,
        )[0]

        kpis.update(quality_kpis)

        # Calculate derived metrics
        kpis["completion_rate"] = (
            flt(kpis["completed_orders"] / kpis["total_orders"] * 100, 2)
            if kpis["total_orders"] > 0
            else 0
        )
        kpis["cancellation_rate"] = (
            flt(kpis["cancelled_orders"] / kpis["total_orders"] * 100, 2)
            if kpis["total_orders"] > 0
            else 0
        )
        kpis["quality_pass_rate"] = (
            flt(kpis["quality_pass_rate"] * 100, 2) if kpis["quality_pass_rate"] else 0
        )

        return kpis


# API Functions for Dashboard Integration
@frappe.whitelist()
def get_procurement_dashboard_data(from_date=None, to_date=None):
    """Get all procurement dashboard data in one call"""

    analyzer = ProcurementCostAnalyzer()

    return {
        "supplier_spend": analyzer.get_supplier_spend_analysis(from_date, to_date),
        "cost_breakdown": analyzer.get_cost_breakdown_analysis(from_date, to_date),
        "kpis": analyzer.get_procurement_kpis(from_date, to_date),
        "generated_at": get_datetime(),
        "currency": analyzer.currency,
    }


@frappe.whitelist()
def get_supplier_comparison_data(suppliers=None, from_date=None, to_date=None):
    """Get comparative data for multiple suppliers"""

    if isinstance(suppliers, str):
        suppliers = json.loads(suppliers)

    analyzer = ProcurementCostAnalyzer()
    comparison_data = []

    for supplier in suppliers:
        supplier_data = analyzer.get_supplier_performance_scorecard(supplier, from_date, to_date)
        if supplier_data["supplier_scorecards"]:
            comparison_data.append(supplier_data["supplier_scorecards"][0])

    return {
        "comparison_data": comparison_data,
        "period": f"{from_date or 'N/A'} to {to_date or 'N/A'}",
        "currency": analyzer.currency,
    }


@frappe.whitelist()
def export_procurement_analytics(format_type="excel", from_date=None, to_date=None):
    """Export procurement analytics data"""

    analyzer = ProcurementCostAnalyzer()

    # Get all analytics data
    supplier_data = analyzer.get_supplier_spend_analysis(from_date, to_date)
    cost_data = analyzer.get_cost_breakdown_analysis(from_date, to_date)
    price_data = analyzer.get_price_trend_analysis(from_date=from_date, to_date=to_date)

    if format_type == "excel":
        return _export_to_excel(supplier_data, cost_data, price_data)
    elif format_type == "pdf":
        return _export_to_pdf(supplier_data, cost_data, price_data)
    else:
        return {
            "supplier_spend": supplier_data,
            "cost_breakdown": cost_data,
            "price_trends": price_data,
        }


def _export_to_excel(supplier_data, cost_data, price_data):
    """Export analytics data to Excel format"""

    try:
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()

        # Supplier spend sheet
        ws1 = wb.active
        ws1.title = "Supplier Spend Analysis"

        # Headers
        headers = [
            "Supplier",
            "Supplier Name",
            "Total Orders",
            "Total Spend",
            "Avg Order Value",
            "Quality Pass Rate",
        ]
        for col, header in enumerate(headers, 1):
            ws1.cell(row=1, column=col, value=header)

        # Data
        for row, supplier in enumerate(supplier_data["supplier_spend_data"], 2):
            ws1.cell(row=row, column=1, value=supplier["supplier"])
            ws1.cell(row=row, column=2, value=supplier["supplier_name"])
            ws1.cell(row=row, column=3, value=supplier["total_orders"])
            ws1.cell(row=row, column=4, value=supplier["total_spend"])
            ws1.cell(row=row, column=5, value=supplier["avg_order_value"])
            ws1.cell(row=row, column=6, value=supplier["quality_pass_rate"])

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return {
            "file_content": output.getvalue(),
            "filename": f"procurement_analytics_{today()}.xlsx",
            "file_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }

    except Exception as e:
        frappe.log_error(f"Excel export error: {str(e)}")
        return {"error": str(e)}


def _export_to_pdf(supplier_data, cost_data, price_data):
    """Export analytics data to PDF format"""

    try:
        # This would implement PDF generation
        # For now, return a placeholder
        return {
            "message": "PDF export functionality to be implemented",
            "filename": f"procurement_analytics_{today()}.pdf",
            "file_type": "application/pdf",
        }

    except Exception as e:
        frappe.log_error(f"PDF export error: {str(e)}")
        return {"error": str(e)}
